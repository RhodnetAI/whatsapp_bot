"""Sales Bot catalog service — CRUD over the Sales Bot's own ``sales_products``
table. Products are added manually or via Excel upload and stored in Supabase
(no Qdrant/vectorization). This table is independent from the Information Bot's
products and is the single source the WhatsApp shopping flow reads from."""

import logging
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from fastapi import BackgroundTasks, HTTPException, UploadFile
from openpyxl import load_workbook

from app.db.supabase_client import first_row, supabase, supabase_admin
from app.models.schemas import (
    SalesProductCreate,
    SalesProductItem,
    SalesProductsUploadStatusResponse,
    SalesProductUpdate,
)

logger = logging.getLogger("whatsapp")

TABLE = "sales_products"

MAX_EXCEL_ROWS = 500

# Maps normalized Excel column headers to sales_products fields.
_COLUMN_ALIASES = {
    "name": "name",
    "description": "description",
    "category": "category",
    "price": "price",
    "currency": "currency",
    "image": "image_url",
    "image url": "image_url",
    "images": "image_url",
    "stock": "stock_quantity",
    "stock quantity": "stock_quantity",
    "status": "status",
    "active": "status",
    "discount": "discount_percentage",
    "discount percentage": "discount_percentage",
    "discount percent": "discount_percentage",
    "discount %": "discount_percentage",
    "google product category": "google_product_category",
    "google category": "google_product_category",
    "size": "size",
    "model": "model",
    "color": "color",
    "colour": "color",
    "delivery date": "delivery_date",
    "quantity to sell": "quantity_to_sell",
    "rating average": "rating_average",
    "rating avg": "rating_average",
    "rating count": "rating_count",
}

# In-memory Excel upload job tracking (admin-only, low volume) — mirrors the
# Information Bot products uploader so the UI can poll per-row progress.
_upload_jobs: dict[str, dict[str, Any]] = {}


def _db():
    return supabase_admin if supabase_admin is not None else supabase


def _generate_retailer_id() -> str:
    return f"sku-{uuid.uuid4().hex[:10]}"


def parse_price_to_minor(value: Any) -> int:
    """Best-effort parse of a free-text price (e.g. '₹1,299.00', '999/mo') into
    integer minor units (paise). Returns 0 when no number is found."""
    if value is None:
        return 0
    match = re.search(r"\d[\d,]*(?:\.\d+)?", str(value))
    if not match:
        return 0
    number = match.group(0).replace(",", "")
    try:
        return int(round(float(number) * 100))
    except ValueError:
        return 0


def _clamp_discount_percentage(value: Any) -> float:
    try:
        v = float(value or 0)
    except (TypeError, ValueError):
        v = 0.0
    return max(0.0, min(100.0, v))


def compute_sale_price_minor(price_minor: int, discount_percentage: float) -> int:
    """The discounted price (in minor units) after applying ``discount_percentage``
    to ``price_minor``. With no discount, the sale price equals the price."""
    if discount_percentage <= 0:
        return int(price_minor)
    return max(0, round(price_minor * (1 - discount_percentage / 100)))


def _require_google_product_category(value: str | None) -> str:
    category = (value or "").strip()
    if not category:
        raise HTTPException(status_code=400, detail="Google Product Category is required")
    return category


def _parse_optional_int(value: Any) -> int | None:
    s = (str(value) if value is not None else "").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _parse_optional_float(value: Any) -> float | None:
    s = (str(value) if value is not None else "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(value: Any) -> str | None:
    s = (str(value) if value is not None else "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _row_to_item(row: dict[str, Any]) -> SalesProductItem:
    stock = row.get("stock_quantity")
    qty_to_sell = row.get("quantity_to_sell")
    rating_avg = row.get("rating_average")
    rating_count = row.get("rating_count")
    return SalesProductItem(
        id=str(row.get("id")),
        retailer_id=row.get("retailer_id") or "",
        name=row.get("name") or "",
        description=row.get("description") or "",
        category=row.get("category") or "",
        price_minor=int(row.get("price_minor") or 0),
        currency=row.get("currency") or "INR",
        image_url=row.get("image_url") or "",
        stock_quantity=int(stock) if stock is not None else None,
        is_active=bool(row.get("is_active", True)),
        meta_catalog_id=row.get("meta_catalog_id"),
        sync_status=row.get("sync_status") or "pending",
        sync_error=row.get("sync_error") or "",
        source=row.get("source") or "manual",
        discount_percentage=float(row.get("discount_percentage") or 0),
        sale_price_minor=int(row.get("sale_price_minor") or 0),
        google_product_category=row.get("google_product_category") or "",
        size=row.get("size") or "",
        model=row.get("model") or "",
        color=row.get("color") or "",
        delivery_date=row.get("delivery_date"),
        quantity_to_sell=int(qty_to_sell) if qty_to_sell is not None else None,
        rating_average=float(rating_avg) if rating_avg is not None else None,
        rating_count=int(rating_count) if rating_count is not None else None,
        created_at=row.get("created_at"),
        updated_at=row.get("updated_at"),
    )


def list_products() -> list[SalesProductItem]:
    result = _db().table(TABLE).select("*").order("created_at", desc=True).execute()
    rows = getattr(result, "data", None) or []
    return [_row_to_item(row) for row in rows if isinstance(row, dict)]


def fetch_active_rows() -> list[dict[str, Any]]:
    """Raw active product rows for the conversational flow."""
    result = (
        _db().table(TABLE).select("*").eq("is_active", True).order("category").execute()
    )
    return [row for row in (getattr(result, "data", None) or []) if isinstance(row, dict)]


def fetch_rows_by_retailer_ids(retailer_ids: list[str]) -> dict[str, dict[str, Any]]:
    """Map retailer_id -> product row, for the given retailer ids."""
    if not retailer_ids:
        return {}
    result = _db().table(TABLE).select("*").in_("retailer_id", retailer_ids).execute()
    out: dict[str, dict[str, Any]] = {}
    for row in getattr(result, "data", None) or []:
        if isinstance(row, dict) and row.get("retailer_id"):
            out[row["retailer_id"]] = row
    return out


def get_row(product_id: str) -> dict[str, Any] | None:
    return first_row(_db().table(TABLE).select("*").eq("id", product_id).execute())


def _insert_product_fields(fields: dict[str, Any], source: str) -> SalesProductItem:
    """Insert one product row from a plain fields dict (shared by manual create
    and Excel upload)."""
    now = datetime.now(timezone.utc).isoformat()
    price_minor = int(fields.get("price_minor") or 0)
    discount_percentage = _clamp_discount_percentage(fields.get("discount_percentage"))
    row = {
        "id": str(uuid.uuid4()),
        "retailer_id": (str(fields.get("retailer_id") or "")).strip() or _generate_retailer_id(),
        "name": fields.get("name") or "",
        "description": fields.get("description") or "",
        "category": fields.get("category") or "",
        "price_minor": price_minor,
        "currency": fields.get("currency") or "INR",
        "image_url": fields.get("image_url") or "",
        "stock_quantity": fields.get("stock_quantity"),
        "is_active": bool(fields.get("is_active", True)),
        "source": source,
        "sync_status": "pending",
        "discount_percentage": discount_percentage,
        "sale_price_minor": compute_sale_price_minor(price_minor, discount_percentage),
        "google_product_category": _require_google_product_category(fields.get("google_product_category")),
        "size": fields.get("size") or "",
        "model": fields.get("model") or "",
        "color": fields.get("color") or "",
        "delivery_date": fields.get("delivery_date") or None,
        "quantity_to_sell": fields.get("quantity_to_sell"),
        "rating_average": fields.get("rating_average"),
        "rating_count": fields.get("rating_count"),
        "created_at": now,
        "updated_at": now,
    }
    try:
        _db().table(TABLE).insert(row).execute()
    except Exception as exc:
        logger.exception("Failed to create sales product")
        raise HTTPException(status_code=400, detail=f"Could not save product: {exc}") from exc
    return _row_to_item(row)


def create_product(payload: SalesProductCreate) -> SalesProductItem:
    return _insert_product_fields(
        {
            "retailer_id": payload.retailer_id,
            "name": payload.name,
            "description": payload.description,
            "category": payload.category,
            "price_minor": int(payload.price_minor or 0),
            "currency": payload.currency or "INR",
            "image_url": payload.image_url,
            "stock_quantity": payload.stock_quantity,
            "is_active": bool(payload.is_active),
            "discount_percentage": payload.discount_percentage,
            "google_product_category": payload.google_product_category,
            "size": payload.size,
            "model": payload.model,
            "color": payload.color,
            "delivery_date": payload.delivery_date,
            "quantity_to_sell": payload.quantity_to_sell,
            "rating_average": payload.rating_average,
            "rating_count": payload.rating_count,
        },
        source="manual",
    )


def update_product(product_id: str, payload: SalesProductUpdate) -> SalesProductItem:
    existing = get_row(product_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="Product not found")
    price_minor = int(payload.price_minor or 0)
    discount_percentage = _clamp_discount_percentage(payload.discount_percentage)
    update = {
        "name": payload.name,
        "description": payload.description,
        "category": payload.category,
        "price_minor": price_minor,
        "currency": payload.currency or "INR",
        "image_url": payload.image_url,
        "stock_quantity": payload.stock_quantity,
        "is_active": bool(payload.is_active),
        "discount_percentage": discount_percentage,
        "sale_price_minor": compute_sale_price_minor(price_minor, discount_percentage),
        "google_product_category": _require_google_product_category(payload.google_product_category),
        "size": payload.size,
        "model": payload.model,
        "color": payload.color,
        "delivery_date": payload.delivery_date or None,
        "quantity_to_sell": payload.quantity_to_sell,
        "rating_average": payload.rating_average,
        "rating_count": payload.rating_count,
        # Any edit means the Meta catalog copy is stale again.
        "sync_status": "pending" if existing.get("sync_status") == "synced" else existing.get("sync_status"),
    }
    _db().table(TABLE).update(update).eq("id", product_id).execute()
    return _row_to_item({**existing, **update})


def delete_product(product_id: str) -> None:
    existing = get_row(product_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="Product not found")
    _db().table(TABLE).delete().eq("id", product_id).execute()


# ── Excel upload ─────────────────────────────────────────────────────────────
def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _excel_row_to_fields(raw: dict[str, str]) -> dict[str, Any]:
    """Convert a parsed Excel row (header->value) into product insert fields."""
    status = (raw.get("status") or "").strip().lower()
    is_active = status not in ("out of stock", "inactive", "no", "false", "0", "disabled")

    stock_raw = (raw.get("stock_quantity") or "").strip()
    stock: int | None = None
    if stock_raw:
        try:
            stock = int(float(stock_raw))
        except ValueError:
            stock = None

    return {
        "name": (raw.get("name") or "").strip(),
        "description": raw.get("description") or "",
        "category": raw.get("category") or "",
        "price_minor": parse_price_to_minor(raw.get("price")),
        "currency": (raw.get("currency") or "INR").strip().upper() or "INR",
        "image_url": (raw.get("image_url") or "").strip(),
        "stock_quantity": stock,
        "is_active": is_active,
        "discount_percentage": _clamp_discount_percentage(_parse_optional_float(raw.get("discount_percentage"))),
        "google_product_category": (raw.get("google_product_category") or "").strip(),
        "size": (raw.get("size") or "").strip(),
        "model": (raw.get("model") or "").strip(),
        "color": (raw.get("color") or "").strip(),
        "delivery_date": _parse_date(raw.get("delivery_date")),
        "quantity_to_sell": _parse_optional_int(raw.get("quantity_to_sell")),
        "rating_average": _parse_optional_float(raw.get("rating_average")),
        "rating_count": _parse_optional_int(raw.get("rating_count")),
    }


def _parse_excel_rows(raw_bytes: bytes) -> list[dict[str, Any]]:
    try:
        workbook: Any = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    column_map: dict[int, str] = {}
    for index, header in enumerate(header_row or []):
        field = _COLUMN_ALIASES.get(_normalize_header(header))
        if field:
            column_map[index] = field

    if "name" not in column_map.values():
        raise HTTPException(status_code=400, detail="Excel file must include a 'Name' column")
    if "google_product_category" not in column_map.values():
        raise HTTPException(status_code=400, detail="Excel file must include a 'Google Product Category' column")

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if raw_row is None or all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        raw_fields: dict[str, str] = {}
        for index, field_name in column_map.items():
            if index < len(raw_row) and raw_row[index] is not None:
                raw_fields[field_name] = str(raw_row[index]).strip()
        fields = _excel_row_to_fields(raw_fields)
        if not fields["name"]:
            continue
        # Google Product Category is mandatory — reject rows missing it.
        if not fields["google_product_category"]:
            continue
        rows.append(fields)
        if len(rows) >= MAX_EXCEL_ROWS:
            break

    if not rows:
        raise HTTPException(status_code=400, detail="No usable rows were found in the Excel file")
    return rows


async def start_excel_upload(file: UploadFile, background_tasks: BackgroundTasks) -> tuple[str, int]:
    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="The uploaded file is empty")
    rows = _parse_excel_rows(raw_bytes)

    job_id = str(uuid.uuid4())
    _upload_jobs[job_id] = {
        "status": "processing",
        "total": len(rows),
        "processed": 0,
        "items": [],
        "error": None,
    }
    background_tasks.add_task(_process_upload_job, job_id, rows)
    return job_id, len(rows)


def _process_upload_job(job_id: str, rows: list[dict[str, Any]]) -> None:
    job = _upload_jobs[job_id]
    try:
        for fields in rows:
            item = _insert_product_fields(fields, source="excel")
            job["items"].append(item)
            job["processed"] += 1
        job["status"] = "done"
    except Exception as exc:
        logger.exception("Sales Excel upload job %s failed", job_id)
        job["status"] = "failed"
        job["error"] = str(exc)


def get_upload_job_status(job_id: str) -> SalesProductsUploadStatusResponse:
    job = _upload_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Upload job not found")
    return SalesProductsUploadStatusResponse(
        status=job["status"],
        total=job["total"],
        processed=job["processed"],
        items=list(job["items"]),
        error=job["error"],
    )


def distinct_active_categories(rows: list[dict[str, Any]]) -> list[str]:
    seen: list[str] = []
    for row in rows:
        cat = (row.get("category") or "").strip() or "Other"
        if cat not in seen:
            seen.append(cat)
    return seen


# ── Read helpers for the "Open Store" data-exchange Flow ─────────────────────
# Thin wrappers over the active-rows query above, returning plain dicts shaped
# for the Flow screens. Used by ``app.api.routes.flow_store`` (Path B); the
# chat-based browsing handlers use ``fetch_active_rows`` / ``get_row`` directly.
def get_categories() -> list[str]:
    """Distinct active categories, in catalog order (``Other`` for blanks)."""
    return distinct_active_categories(fetch_active_rows())


def get_products_by_category(category: str) -> list[dict[str, Any]]:
    """All active product rows whose category matches ``category`` (blank
    categories fold into ``Other``, matching :func:`get_categories`)."""
    target = (category or "").strip() or "Other"
    out: list[dict[str, Any]] = []
    for row in fetch_active_rows():
        cat = (row.get("category") or "").strip() or "Other"
        if cat == target:
            out.append(row)
    return out


def get_product(product_id: str) -> dict[str, Any] | None:
    """A single active product row by id, or None if missing/inactive."""
    row = get_row(product_id)
    if not row or not row.get("is_active", True):
        return None
    return row
