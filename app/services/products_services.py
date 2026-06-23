import logging
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Literal, cast

from fastapi import BackgroundTasks, HTTPException, UploadFile
from openpyxl import load_workbook

from app.db.supabase_client import first_row, supabase, supabase_admin
from app.models.schemas import (
    ProductServiceCreate,
    ProductServiceItem,
    ProductServiceUpdate,
    ProductsServicesUploadStatusResponse,
)
from app.services.products_vectorizer import delete_product_vector, store_product_vector

logger = logging.getLogger("whatsapp")

TABLE_NAME = "information_bot_products_services"

FIELD_NAMES = [
    "name",
    "short_description",
    "full_description",
    "category",
    "price",
    "discount_price",
    "status",
    "images",
    "rating",
    "reviews_count",
    "purchased_count",
]

# Maps normalized Excel column headers to the field they populate.
_COLUMN_ALIASES = {
    "name": "name",
    "short description": "short_description",
    "full description": "full_description",
    "category": "category",
    "price": "price",
    "discount price": "discount_price",
    "status": "status",
    "images": "images",
    "rating": "rating",
    "reviews count": "reviews_count",
    "purchased count": "purchased_count",
}

MAX_EXCEL_ROWS = 500

# In-memory tracking for Excel upload jobs so the frontend can poll real
# per-row progress. This is an admin-only, low-volume settings tool (mirrors
# the module-level client caching already used in vectorizer.py), so a plain
# dict is sufficient — no need for a database-backed job table.
_upload_jobs: dict[str, dict[str, Any]] = {}


def _db():
    """Return the admin client (bypasses RLS) when available, else fall back to anon client."""
    return supabase_admin if supabase_admin is not None else supabase


def _row_to_item(row: dict[str, Any]) -> ProductServiceItem:
    return ProductServiceItem(
        id=str(row.get("id")),
        kind=cast(Literal["product", "service"], row.get("kind") or "product"),
        name=row.get("name") or "",
        short_description=row.get("short_description") or "",
        full_description=row.get("full_description") or "",
        category=row.get("category") or "",
        price=row.get("price") or "",
        discount_price=row.get("discount_price") or "",
        status=row.get("status") or "Active",
        images=row.get("images") or "",
        rating=row.get("rating") or "",
        reviews_count=row.get("reviews_count") or "",
        purchased_count=row.get("purchased_count") or "",
        source=row.get("source") or "manual",
        vectorization_status=row.get("vectorization_status") or "processing",
        created_at=row.get("created_at"),
        updated_at=row.get("updated_at"),
    )


def list_items() -> list[ProductServiceItem]:
    result = _db().table(TABLE_NAME).select("*").order("created_at", desc=True).execute()
    rows = getattr(result, "data", None) or []
    return [_row_to_item(row) for row in rows]


async def _create_row(kind: str, fields: dict[str, str], source: str) -> ProductServiceItem:
    item_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    row = {
        "id": item_id,
        "kind": kind,
        **fields,
        "source": source,
        "vectorization_status": "processing",
        "created_at": now,
        "updated_at": now,
    }
    _db().table(TABLE_NAME).insert(row).execute()

    return _row_to_item(row)


async def _vectorize_item(item_id: str, fields: dict[str, Any]) -> str:
    """Embed a product/service row and flip vectorization_status to done/failed
    — run as a background task so create/update requests aren't blocked on an
    embeddings API call. Returns the resulting status."""
    try:
        await store_product_vector(item_id, fields)
        status = "done"
    except Exception:
        logger.exception("Failed to vectorize product/service %s", item_id)
        status = "failed"
    _db().table(TABLE_NAME).update({"vectorization_status": status}).eq("id", item_id).execute()
    return status


async def create_item(payload: ProductServiceCreate, background_tasks: BackgroundTasks) -> ProductServiceItem:
    fields = payload.model_dump(exclude={"kind"})
    item = await _create_row(payload.kind, fields, source="manual")
    background_tasks.add_task(_vectorize_item, item.id, fields)
    return item


async def update_item(
    item_id: str, payload: ProductServiceUpdate, background_tasks: BackgroundTasks
) -> ProductServiceItem:
    existing = first_row(_db().table(TABLE_NAME).select("*").eq("id", item_id).execute())
    if existing is None:
        raise HTTPException(status_code=404, detail="Product/service not found")

    fields = payload.model_dump()
    now = datetime.now(timezone.utc).isoformat()
    update_payload = {**fields, "vectorization_status": "processing", "updated_at": now}
    _db().table(TABLE_NAME).update(update_payload).eq("id", item_id).execute()
    background_tasks.add_task(_vectorize_item, item_id, fields)

    return _row_to_item({**existing, **update_payload})


async def delete_item(item_id: str) -> None:
    existing = first_row(_db().table(TABLE_NAME).select("*").eq("id", item_id).execute())
    if existing is None:
        raise HTTPException(status_code=404, detail="Product/service not found")

    _db().table(TABLE_NAME).delete().eq("id", item_id).execute()
    try:
        await delete_product_vector(item_id)
    except Exception:
        logger.warning("Failed to delete vector for product/service %s", item_id)


async def reindex_all(background_tasks: BackgroundTasks) -> int:
    """Backfill vectors for every existing row. Needed once after enabling
    semantic retrieval: rows created before this change have
    vectorization_status='done' (the old hardcoded value) but were never
    actually embedded, so search_products() would find nothing for them
    until they're individually edited or this is run."""
    result = _db().table(TABLE_NAME).select("id, " + ", ".join(FIELD_NAMES)).execute()
    rows = [row for row in (result.data or []) if isinstance(row, dict)]

    for row in rows:
        item_id = str(row.get("id"))
        fields = {name: row.get(name) or "" for name in FIELD_NAMES}
        _db().table(TABLE_NAME).update({"vectorization_status": "processing"}).eq("id", item_id).execute()
        background_tasks.add_task(_vectorize_item, item_id, fields)

    return len(rows)


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _parse_excel_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        workbook: Any = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
        if workbook is None:
            raise HTTPException(status_code=400, detail="Could not read Excel file")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    column_map: dict[int, str] = {}
    for index, header in enumerate(header_row or []):
        field = _COLUMN_ALIASES.get(_normalize_header(header))
        if field:
            column_map[index] = field

    if "name" not in column_map.values():
        raise HTTPException(status_code=400, detail="Excel file must include a 'Name' column")

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        if raw_row is None or all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue

        fields = {field_name: "" for field_name in FIELD_NAMES}
        for index, field_name in column_map.items():
            if index < len(raw_row) and raw_row[index] is not None:
                fields[field_name] = str(raw_row[index]).strip()

        if not fields["name"]:
            continue

        rows.append(fields)
        if len(rows) >= MAX_EXCEL_ROWS:
            break

    if not rows:
        raise HTTPException(status_code=400, detail="No usable rows were found in the Excel file")

    return rows


async def start_excel_upload(kind: str, file: UploadFile, background_tasks: BackgroundTasks) -> tuple[str, int]:
    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="The uploaded file is empty")

    rows = _parse_excel_rows(raw_bytes)

    job_id = str(uuid.uuid4())
    _upload_jobs[job_id] = {
        "status": "processing",
        "total": len(rows),
        "processed": 0,
        "items": [],
        "error": None,
    }
    background_tasks.add_task(_process_upload_job, job_id, kind, rows)
    return job_id, len(rows)


async def _process_upload_job(job_id: str, kind: str, rows: list[dict[str, str]]) -> None:
    job = _upload_jobs[job_id]
    try:
        for fields in rows:
            item = await _create_row(kind, fields, source="excel")
            item.vectorization_status = await _vectorize_item(item.id, fields)
            job["items"].append(item)
            job["processed"] += 1
        job["status"] = "done"
    except Exception as exc:
        logger.exception("Excel upload job %s failed", job_id)
        job["status"] = "failed"
        job["error"] = str(exc)


def get_upload_job_status(job_id: str) -> ProductsServicesUploadStatusResponse:
    job = _upload_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Upload job not found")

    return ProductsServicesUploadStatusResponse(
        status=job["status"],
        total=job["total"],
        processed=job["processed"],
        items=list(job["items"]),
        error=job["error"],
    )
